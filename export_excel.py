"""posts.db → Excel 내보내기.

사용법:
  .venv/bin/python export_excel.py
  .venv/bin/python export_excel.py --out data/my_report.xlsx
"""
import argparse
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DB_PATH   = Path(__file__).parent / "data" / "posts.db"
OUT_DIR   = Path(__file__).parent / "data"


def export(out_path: Path, since: str | None = None):
    conn = sqlite3.connect(DB_PATH)

    # since 가 없으면 72시간 이내 전체, 있으면 해당 수집 시점 이후 스냅샷이 기록된 게시물만
    if since:
        time_filter_posts    = "INNER JOIN post_snapshots ps ON p.media_id = ps.media_id"
        time_where_posts     = "AND ps.collected_at >= :since AND p.media_timestamp >= datetime('now', '-72 hours')"
        time_filter_summary  = "INNER JOIN post_snapshots ps ON p.media_id = ps.media_id"
        time_where_summary   = "AND ps.collected_at >= :since AND p.media_timestamp >= datetime('now', '-72 hours')"
        params = {"since": since}
    else:
        time_filter_posts    = ""
        time_where_posts     = "AND p.media_timestamp >= datetime('now', '-72 hours')"
        time_filter_summary  = ""
        time_where_summary   = "AND p.media_timestamp >= datetime('now', '-72 hours')"
        params = {}

    # ── 시트 1: 게시물 전체 ──────────────────────────────
    posts = pd.read_sql_query(f"""
        SELECT DISTINCT
            p.account_name                          AS 계정명,
            COALESCE(fc.follower_count, 0)          AS 팔로워수,
            p.media_type                            AS 미디어타입,
            p.like_count                            AS 좋아요,
            p.comment_count                         AS 댓글수,
            p.caption                               AS 캡션,
            p.permalink                             AS 게시물링크,
            p.media_timestamp                       AS 게시일시,
            p.first_seen_at                         AS 최초수집일시
        FROM posts p
        {time_filter_posts}
        LEFT JOIN (
            SELECT account_name, MAX(follower_count) AS follower_count
            FROM posts
            WHERE follower_count > 0
            GROUP BY account_name
        ) fc ON p.account_name = fc.account_name
        WHERE p.comment_count >= 300
          {time_where_posts}
        ORDER BY p.media_timestamp DESC
    """, conn, params=params)

    # 날짜 포맷 정리 (UTC → KST 표시)
    for col in ["게시일시", "최초수집일시"]:
        posts[col] = pd.to_datetime(posts[col], utc=True).dt.tz_convert("Asia/Seoul").dt.strftime("%Y-%m-%d %H:%M")

    posts["팔로워수"] = posts["팔로워수"].fillna(0).astype(int)
    posts["좋아요"]   = posts["좋아요"].fillna(0).astype(int)
    posts["댓글수"]   = posts["댓글수"].fillna(0).astype(int)

    # ── 시트 2: 계정별 요약 ──────────────────────────────
    summary = pd.read_sql_query(f"""
        SELECT
            p.account_name                      AS 계정명,
            COALESCE(fc.follower_count, 0)      AS 팔로워수,
            COUNT(DISTINCT p.media_id)          AS 게시물수,
            SUM(p.like_count)                   AS 총좋아요,
            SUM(p.comment_count)                AS 총댓글수,
            ROUND(AVG(p.like_count), 1)         AS 평균좋아요,
            ROUND(AVG(p.comment_count), 1)      AS 평균댓글수,
            MAX(p.media_timestamp)              AS 최신게시일시
        FROM posts p
        {time_filter_summary}
        LEFT JOIN (
            SELECT account_name, MAX(follower_count) AS follower_count
            FROM posts
            WHERE follower_count > 0
            GROUP BY account_name
        ) fc ON p.account_name = fc.account_name
        WHERE p.account_name != ''
          AND p.comment_count >= 300
          {time_where_summary}
        GROUP BY p.account_name
        ORDER BY 총댓글수 DESC
    """, conn, params=params)

    summary["최신게시일시"] = pd.to_datetime(summary["최신게시일시"], utc=True).dt.tz_convert("Asia/Seoul").dt.strftime("%Y-%m-%d %H:%M")
    summary["팔로워수"]     = summary["팔로워수"].fillna(0).astype(int)

    conn.close()

    # ── Excel 저장 ───────────────────────────────────────
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        posts.to_excel(writer,   sheet_name="게시물 전체", index=False)
        summary.to_excel(writer, sheet_name="계정별 요약", index=False)

        # 게시물링크 하이퍼링크 적용
        ws_posts = writer.sheets["게시물 전체"]
        link_col  = posts.columns.get_loc("게시물링크") + 1
        link_font = Font(color="0563C1", underline="single")
        for row_idx, url in enumerate(posts["게시물링크"], start=2):
            if url and isinstance(url, str) and url.startswith("http"):
                cell = ws_posts.cell(row=row_idx, column=link_col)
                cell.hyperlink = url
                cell.value     = url
                cell.font      = link_font

        # 열 너비 자동 조정
        for sheet_name, df in [("게시물 전체", posts), ("계정별 요약", summary)]:
            ws = writer.sheets[sheet_name]
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).str.len().max(), len(col)) + 2
                ws.column_dimensions[get_column_letter(i)].width = min(max_len, 50)

    print(f"저장 완료: {out_path}")
    print(f"  게시물 전체: {len(posts):,}행")
    print(f"  계정별 요약: {len(summary):,}행")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=None)
    parser.add_argument("--since", type=str, default=None, help="이 시각(ISO) 이후 수집된 게시물만 추출")
    args = parser.parse_args()

    now_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    out_path = args.out or OUT_DIR / f"벤치시트_{now_str}.xlsx"
    export(out_path, since=args.since)
